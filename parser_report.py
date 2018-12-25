import os
import json
import openpyxl
import time

class ExcelParser:
    FIRST_TABLE = None
    NOT_ISRAEL_WORDS = ['מט"ח', 'חוץ לארץ', 'חו"ל']
    ISRAEL_WORDS = ['ישראל', 'בארץ']
    FIRST_FIELD_TABLE = ['שם נ"ע', 'שם המנפיק/שם נייר ערך']

    def __init__(self, logger):
        self._logger = logger
        self._is_israel = None

    def parse_file(self, file_path):
        """
        Get pension report excel file and parse data by sheet
        Move over all excel data sheet and parse
        :param file_path: full file path
        :return: parsed data :type: dictionary
            """
        # Load in the workbook file
        workbook = openpyxl.load_workbook(file_path)

        if not workbook:
            self._logger.error("Failed to load excel file")
            return False

        # Move over the all sheets
        for sheet_name in workbook.sheetnames:
            if workbook[sheet_name].title == 'סכום נכסי הקרן':
                # :todo: need parse this sheet ?
                continue
            # Parse sheet
            if not self._parse_sheet(sheet=workbook[sheet_name]):
                self._logger.error("Failed to parser sheet {0}".format(sheet_name))

    def _parse_sheet(self, sheet, start_row=0, start_column=2):
        """
        Parse excel pension report sheet
        :param sheet: openpyxl sheet obj
        :param start_row: row number to start
        :param start_column: column numbrt to start
        :return: True / False
        """
        current_row = start_row
        current_column = start_column
        current_cell = None
        row_data = None

        data = {
            "metadata": {
                "אפיק השקעה": sheet.title

            },
            "data": {}
            }

        # Parse metadata, stop when find the first table field
        while current_cell not in self.FIRST_FIELD_TABLE:
            if current_cell:
                metadata = self._get_metadata(data=row_data)
                if metadata:
                    data["metadata"][metadata[0]] = metadata[1]

            current_row += 1
            row_data = ExcelAdapter.get_entire_row(sheet=sheet,
                                                   row=current_row,
                                                   min_column=current_column)

            if row_data:
                # strip all spaces from start and end string
                current_cell = row_data[0].strip()
            else:
                current_cell = None
        else:
            # Get fields name
            fields_name = ExcelAdapter.get_entire_row(sheet=sheet,
                                                      row=current_row,
                                                      min_column=start_column)

            fields_len = len(fields_name)

        empty_len = 0
        current_cell = ""
        while current_cell != '* בעל ענין/צד קשור':

            if empty_len > 5:
                # self._logger.warn("max empty row")
                break

            # Get next row
            current_row += 1
            data_row = ExcelAdapter.get_entire_row(sheet=sheet,
                                                   row=current_row,
                                                   min_column=start_column,
                                                   max_column=fields_len+start_column)

            # Check if is empty row or first cell is empty
            if not data_row or not data_row[0]:
                empty_len += 1
                continue

            current_cell = data_row[0]

            # if current cell start is total row
            if current_cell.find('סה"כ') != -1:
                self._parse_total_field(current_cell)
            else:
                data["data"][data_row[0]] = {
                    'שייכות למדד': self._total_data,
                    "israel": self._is_israel
                }
                for i in range(1, fields_len):
                    try:
                        data["data"][data_row[0]][fields_name[i]] = data_row[i]
                    except IndexError as ex:
                        self._logger.error("Failed {0} {1}".format(ex, fields_name))

        # print(data)
        path = '/tmp'
        file_name = "{0}.json".format(data["metadata"]['אפיק השקעה'])
        self._save_to_json_file(path=path, file_name=file_name, data=data)
        return True

    def _get_metadata(self, data):
        """
        Parse metadata data
        :param data: list of data
        :return:
        """
        first_cell = data[0]
        if not first_cell:
            self._logger.error("No data in first cell")
            return None, None

        finder = first_cell.find(":")
        # find func return -1 when not find
        if finder != -1:
            # Check if the colon char is not last data char (The meaning data in the first cell)
            if len(first_cell) > finder:
                return first_cell[:finder], first_cell[finder+1:]

        # check if len of data is bigger than one
        elif len(data) > 1:
            return first_cell, data[1]

    def _parse_total_field(self, data):
        """
        Parse total field, total filed start with 'סה"כ' word
        In total field we get if the investment in Israel
        :param data:
        :return:
        """
        # strip 'סה"כ' word
        self._total_data = data.strip('סה"כ')

        # lambda warp for string find function
        # Finder return True/False , instead of number in find function
        # Check if search word in self._total_data
        finder = lambda search_word: False if self._total_data.find(search_word) == -1 else True

        # Get words list and use finder lambda and filter function to check if one or more of word_list in data string
        recursive_finder = lambda words_list: True if len(list(filter(finder, words_list))) else False

        if recursive_finder(words_list=self.ISRAEL_WORDS):
            self._is_israel = True
            # print("{0} is israel".format(data))
        elif recursive_finder(words_list=self.NOT_ISRAEL_WORDS):
            self._is_israel = False
            # print("{0} is not israel".format(data))

        # print(self._total_data)

    def _save_to_json_file(self, path, file_name, data):
        if not os.path.isdir(path):
            self._logger.error("folder not exists {0}".format(path))

        full_path = os.path.join(path,file_name)
        try:
            with open(full_path, "w") as outfile:
                json.dump(data, outfile)
            return True
        except Exception as ex:
            self._logger.error("Failed to write json file {0}".format(ex))
            return False


class ExcelAdapter:
    @staticmethod
    def get_cell(sheet, row, column):
        """
        Get cell value
        :param sheet:
        :param row:
        :param column:
        :return:
        """
        try:
            return sheet.cell(row=row, column=column).value
        except Exception as ex:
            raise Exception("Failed to read cell {0}".format(ex))

    @staticmethod
    def get_entire_row(sheet, row, min_column=1, max_column=None):
        """
        Get row between min column number to max column number
        if max column is None, get all cells until cell data is None
        :param sheet:
        :param row:
        :param min_column:
        :return: row data :type: list
        """
        cell_data = None
        row_data = []
        column = min_column

        # lambdas function
        data_exists = lambda: True if cell_data else False
        is_not_max_column = lambda: not(max_column == column)

        # If max column than use is_not_max_column lambda to check if is the max column
        # If max column is None, use data_exists lambda to check if cell data exists
        if max_column:
            check = is_not_max_column
        else:
            check = data_exists

        # Get cell data
        cell_data = ExcelAdapter.get_cell(sheet=sheet, column=column, row=row)

        while check():
            row_data.append(cell_data)
            column += 1
            cell_data = ExcelAdapter.get_cell(sheet=sheet, column=column, row=row)

        return row_data


class FakeLogger:
    def error(self, msg):
        print("error {0}".format(msg))

    def info(self, msg):
        print("info {0}".format(msg))

    def warn(self,msg):
        print("warring {0}".format(msg))

if __name__ == '__main__':
    logger = FakeLogger()
    process_xl = ExcelParser(logger=logger)
    process_xl.parse_file(file_path="test.xlsx")
